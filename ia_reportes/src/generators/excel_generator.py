from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from src.config.settings import REPORTS_OUTPUT_DIR, PANADERIA

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Generador de reportes en Excel"""
    
    def __init__(self):
        self.color_header = "1a472a"
        self.color_alt = "e8f4f0"
    
    def generar_reporte_ventas(self, datos: dict) -> str:
        """Genera reporte de ventas en Excel"""
        try:
            filename = f"{REPORTS_OUTPUT_DIR}/Reporte_Ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb = Workbook()
            
            # Hoja 1: Resumen
            ws_resumen = wb.active
            ws_resumen.title = "Resumen"
            self._crear_resumen_ventas(ws_resumen, datos)
            
            # Hoja 2: Detalle Productos
            if datos.get('productos_vendidos_detalle'):
                ws_productos = wb.create_sheet("Productos")
                self._crear_detalle_productos(ws_productos, datos.get('productos_vendidos_detalle'))
            
            # Hoja 3: Detalle Clientes
            if datos.get('clientes_detalle'):
                ws_clientes = wb.create_sheet("Clientes")
                self._crear_detalle_clientes(ws_clientes, datos.get('clientes_detalle'))
            
            # Hoja 4: Por Categoría
            if datos.get('por_categoria'):
                ws_categoria = wb.create_sheet("Por Categoría")
                self._crear_detalle_categoria(ws_categoria, datos.get('por_categoria'))
            
            wb.save(filename)
            logger.info(f"Excel de ventas generado: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error generando Excel de ventas: {str(e)}")
            raise
    
    def generar_reporte_inventario(self, datos: dict) -> str:
        """Genera reporte de inventario en Excel"""
        try:
            filename = f"{REPORTS_OUTPUT_DIR}/Reporte_Inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb = Workbook()
            
            ws = wb.active
            ws.title = "Inventario"
            
            # Encabezado
            ws['A1'] = PANADERIA['nombre']
            ws['A2'] = 'REPORTE DE INVENTARIO'
            ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y')}"
            
            # Resumen
            row = 5
            ws[f'A{row}'] = "STOCK TOTAL (Valor):"
            ws[f'B{row}'] = datos.get('stock_total_valor', 0)
            ws[f'B{row}'].number_format = '$#,##0.00'
            
            row += 1
            ws[f'A{row}'] = "Cantidad de Items:"
            ws[f'B{row}'] = datos.get('cantidad_items', 0)
            
            row += 1
            ws[f'A{row}'] = "Items Bajo Stock:"
            ws[f'B{row}'] = datos.get('productos_bajo_stock', 0)
            
            # Detalle de insumos
            if datos.get('insumos_detalle'):
                row = 10
                headers = ['Item', 'Categoría', 'Stock Actual', 'Stock Mínimo', 'Estado']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    self._aplicar_estilo_header(cell)
                
                row += 1
                for insumo in datos.get('insumos_detalle', []):
                    ws.cell(row=row, column=1, value=insumo.get('nombre'))
                    ws.cell(row=row, column=2, value=insumo.get('categoria'))
                    ws.cell(row=row, column=3, value=insumo.get('cantidad'))
                    ws.cell(row=row, column=4, value=insumo.get('cantidad_minima'))
                    ws.cell(row=row, column=5, value=insumo.get('estado'))
                    row += 1
            
            # Auto-ajustar columnas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            wb.save(filename)
            logger.info(f"Excel de inventario generado: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error generando Excel de inventario: {str(e)}")
            raise
    
    def generar_reporte_completo(self, datos_multiples: dict) -> str:
        """Genera Excel con múltiples módulos"""
        try:
            filename = f"{REPORTS_OUTPUT_DIR}/Reporte_Completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb = Workbook()
            wb.remove(wb.active)
            
            # Crear hojas por módulo
            if datos_multiples.get('ventas'):
                ws = wb.create_sheet('Ventas')
                self._crear_resumen_ventas(ws, datos_multiples['ventas'])
            
            if datos_multiples.get('inventario'):
                ws = wb.create_sheet('Inventario')
                self._crear_resumen_inventario(ws, datos_multiples['inventario'])
            
            if datos_multiples.get('compras'):
                ws = wb.create_sheet('Compras')
                self._crear_resumen_compras(ws, datos_multiples['compras'])
            
            if datos_multiples.get('produccion'):
                ws = wb.create_sheet('Producción')
                self._crear_resumen_produccion(ws, datos_multiples['produccion'])
            
            wb.save(filename)
            logger.info(f"Excel completo generado: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error generando Excel completo: {str(e)}")
            raise
    
    def _crear_resumen_ventas(self, ws, datos: dict):
        """Crea hoja resumen de ventas"""
        ws['A1'] = PANADERIA['nombre']
        ws['A2'] = 'RESUMEN DE VENTAS'
        ws['A3'] = f"Período: {datos.get('periodo', 'N/A')}"
        
        row = 5
        metricas = [
            ('Total Ventas', f"${datos.get('total_ventas', 0):,.2f}"),
            ('Cantidad de Pedidos', datos.get('cantidad_ordenes', 0)),
            ('Ticket Promedio', f"${datos.get('ticket_promedio', 0):,.2f}"),
            ('Productos Vendidos', datos.get('productos_vendidos', 0)),
            ('Clientes Únicos', datos.get('clientes_unicos', 0)),
            ('Variación M.A.', f"{datos.get('variacion_mes_anterior', 0):+.2f}%"),
        ]
        
        for metrica, valor in metricas:
            ws[f'A{row}'] = metrica
            ws[f'B{row}'] = valor
            row += 1
    
    def _crear_resumen_inventario(self, ws, datos: dict):
        """Crea hoja resumen de inventario"""
        ws['A1'] = 'ESTADO DE INVENTARIO'
        row = 3
        
        ws[f'A{row}'] = 'Stock Total (Valor):'
        ws[f'B{row}'] = datos.get('stock_total_valor', 0)
        row += 1
        
        ws[f'A{row}'] = 'Cantidad Items:'
        ws[f'B{row}'] = datos.get('cantidad_items', 0)
        row += 1
        
        ws[f'A{row}'] = 'Items Bajo Stock:'
        ws[f'B{row}'] = datos.get('productos_bajo_stock', 0)
    
    def _crear_resumen_compras(self, ws, datos: dict):
        """Crea hoja resumen de compras"""
        ws['A1'] = 'RESUMEN DE COMPRAS'
        row = 3
        
        ws[f'A{row}'] = 'Total Compras:'
        ws[f'B{row}'] = datos.get('total_compras', 0)
        row += 1
        
        ws[f'A{row}'] = 'Cantidad Órdenes:'
        ws[f'B{row}'] = datos.get('cantidad_ordenes_compra', 0)
    
    def _crear_resumen_produccion(self, ws, datos: dict):
        """Crea hoja resumen de producción"""
        ws['A1'] = 'RESUMEN DE PRODUCCIÓN'
        row = 3
        
        ws[f'A{row}'] = 'Total Producido:'
        ws[f'B{row}'] = datos.get('total_producido', 0)
        row += 1
        
        ws[f'A{row}'] = 'Cantidad Lotes:'
        ws[f'B{row}'] = datos.get('cantidad_lotes', 0)
    
    def _crear_detalle_productos(self, ws, productos: list):
        """Crea detalle de productos vendidos"""
        headers = ['Producto', 'Cantidad', 'Precio Promedio', 'Total Vendido']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._aplicar_estilo_header(cell)
        
        for row, prod in enumerate(productos, 2):
            ws.cell(row=row, column=1, value=prod.get('nombre'))
            ws.cell(row=row, column=2, value=prod.get('cantidad'))
            ws.cell(row=row, column=3, value=prod.get('precio_promedio'))
            ws.cell(row=row, column=4, value=prod.get('total'))
    
    def _crear_detalle_clientes(self, ws, clientes: list):
        """Crea detalle de clientes"""
        headers = ['Cliente', 'Pedidos', 'Total Gastado', 'Ticket Promedio', 'Última Compra']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._aplicar_estilo_header(cell)
        
        for row, cliente in enumerate(clientes, 2):
            ws.cell(row=row, column=1, value=cliente.get('nombre'))
            ws.cell(row=row, column=2, value=cliente.get('pedidos'))
            ws.cell(row=row, column=3, value=cliente.get('total'))
            ws.cell(row=row, column=4, value=cliente.get('promedio'))
            ws.cell(row=row, column=5, value=cliente.get('ultima_compra'))
    
    def _crear_detalle_categoria(self, ws, categorias: list):
        """Crea detalle por categoría"""
        headers = ['Categoría', 'Cantidad', 'Total Vendido', 'Pedidos']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._aplicar_estilo_header(cell)
        
        for row, cat in enumerate(categorias, 2):
            ws.cell(row=row, column=1, value=cat.get('nombre'))
            ws.cell(row=row, column=2, value=cat.get('cantidad'))
            ws.cell(row=row, column=3, value=cat.get('total'))
            ws.cell(row=row, column=4, value=cat.get('pedidos'))
    
    def _aplicar_estilo_header(self, cell):
        """Aplica estilo a headers"""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=self.color_header, end_color=self.color_header, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
